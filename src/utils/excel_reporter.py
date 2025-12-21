"""
Excel report generation utilities.
"""
from pathlib import Path
from typing import List
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

from models import FiscalDocument, ServiceItem


class ExcelReporter:
    """Generates formatted Excel reports from fiscal documents"""
    
    # Style definitions
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER_THIN = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def _format_date(self, dt: datetime.date) -> str:
        """Format date as DD/MM/YYYY"""
        if not dt:
            return None
        return dt.strftime("%d/%m/%Y")
    
    def _clean_cnpj(self, value: str) -> str:
        """Remove punctuation from CNPJ/CPF"""
        if not value:
            return None
        return "".join(filter(str.isdigit, str(value)))

    def generate_report(self, documents: List[FiscalDocument]) -> Path:
        """
        Generate Excel report with two sheets:
        1. Documentos Fiscais (one row per document)
        2. Itens e Serviços (multiple rows per document)
        """
        if not documents:
            raise ValueError("No documents to generate report")
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.output_dir / f"relatorio_fiscal_{timestamp}.xlsx"
        
        # Create DataFrames
        df_documents = self._create_documents_dataframe(documents)
        df_items = self._create_items_dataframe(documents)
        
        # Write to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_documents.to_excel(writer, sheet_name='Documentos Fiscais', index=False)
            df_items.to_excel(writer, sheet_name='Itens e Serviços', index=False)
        
        # Apply formatting
        self._apply_formatting(output_file)
        
        logger.info(f"Generated Excel report: {output_file}")
        return output_file
    
    def _create_documents_dataframe(self, documents: List[FiscalDocument]) -> pd.DataFrame:
        """Create DataFrame for documents sheet"""
        rows = []
        
        for doc in documents:
            row = {
                # Columns in exact order matching web version
                'Tipo Documento': doc.document_type.value,
                'Número Documento': doc.numero,
                'Data Emissão': self._format_date(doc.data_emissao),
                'Data Saída/Entrada': self._format_date(doc.data_saida_entrada),
                
                # Emitente
                'Emitente CNPJ/CPF': self._clean_cnpj(doc.emitente.cnpj) if doc.emitente else None,
                'Emitente Nome/Razão Social': doc.emitente.razao_social if doc.emitente else None,
                'Emitente Endereço': doc.emitente.endereco.to_string() if doc.emitente and doc.emitente.endereco else None,
                
                # Destinatário
                'Destinatário CNPJ/CPF': self._clean_cnpj(doc.destinatario.cnpj) if doc.destinatario else None,
                
                # Filiais
                'COLIGADA': doc.coligada,
                'FILIAL': doc.filial,
                
                'Destinatário Nome/Razão Social': doc.destinatario.razao_social if doc.destinatario else None,
                'Destinatário Endereço': doc.destinatario.endereco.to_string() if doc.destinatario and doc.destinatario.endereco else None,
                
                # Valores
                'Valor Total Documento': doc.valores.valor_total if doc.valores else None,
                'Valor Líquido Documento': doc.valores.valor_liquido if doc.valores else None,
                'Valor Total Produtos/Serviços': doc.valores.valor_servicos if doc.valores else None,
                'Valor Frete': None,  # Not currently extracted
                'Valor Desconto': doc.valores.desconto if doc.valores else None,
                
                # Impostos (NF-e: ICMS, IPI, PIS, COFINS)
                'ICMS': doc.valores.icms if doc.valores else None,
                'IPI': doc.valores.ipi if doc.valores else None,
                'PIS': doc.valores.pis if doc.valores else None,
                'COFINS': doc.valores.cofins if doc.valores else None,
                
                # ISS (Devido para NFS-e)
                'ISS': doc.valores.iss if doc.valores else None,
                
                # Retenções (NFS-e: valores retidos na fonte)
                'IRRF Retido': doc.valores.ir if doc.valores else None,
                'INSS Retido': doc.valores.inss if doc.valores else None,
                'PIS Retido': doc.valores.pis_retido if doc.valores else None,
                'COFINS Retido': doc.valores.cofins_retido if doc.valores else None,
                'CSLL Retida': doc.valores.csll_retida if doc.valores else None,
                'ISS Retido (Serviço)': doc.valores.iss_retido if doc.valores else None,
                
                'Chave Acesso NF-e': doc.chave_acesso,
                'Observações Extração': doc.error_message if doc.error_message else ("Documento Escaneado" if doc.is_scanned else None),
            }
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _create_items_dataframe(self, documents: List[FiscalDocument]) -> pd.DataFrame:
        """Create DataFrame for items/services sheet"""
        rows = []
        
        for doc in documents:
            if not doc.itens:
                # Add empty row to maintain document reference
                rows.append({
                    'Arquivo': doc.filename,
                    'Número Documento': doc.numero,
                    'Item': None,
                    'Código': None,
                    'Descrição': None,
                    'Quantidade': None,
                    'Unidade': None,
                    'Valor Unitário': None,
                    'Valor Total': None,
                    'Alíquota ISS (%)': None,
                    'Valor ISS': None,
                })
            else:
                for item in doc.itens:
                    row = {
                        'Arquivo': doc.filename,
                        'Número Documento': doc.numero,
                        'Item': item.item_numero,
                        'Código': item.codigo,
                        'Descrição': item.descricao,
                        'Quantidade': item.quantidade,
                        'Unidade': item.unidade,
                        'Valor Unitário': item.valor_unitario,
                        'Valor Total': item.valor_total,
                        'Alíquota ISS (%)': item.aliquota_iss,
                        'Valor ISS': item.valor_iss,
                    }
                    rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _apply_formatting(self, excel_file: Path):
        """Apply Excel formatting (headers, borders, column widths)"""
        wb = load_workbook(excel_file)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.BORDER_THIN
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Apply filters
            ws.auto_filter.ref = ws.dimensions
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        wb.save(excel_file)
        logger.debug(f"Applied formatting to {excel_file}")
